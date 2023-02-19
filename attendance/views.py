from django.contrib.admin.views.decorators import staff_member_required
from openpyxl import Workbook ,load_workbook
from django.shortcuts import render , redirect
from django.http import response , HttpResponse
from attendance.models import log, Attendance,User
from datetime import datetime,date,time
from django.contrib.auth.decorators import login_required
# import xlwt
import datetime
from openpyxl.styles import Color, Fill
from openpyxl.cell import Cell
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from django.core.paginator import Paginator,PageNotAnInteger,EmptyPage
from django.contrib import messages


# Storing Loggin Logout Time
# def Attendance_store(request):
#     a=[]
#     for i in log.objects.all():
#         a.append(i.user_id.id)
#         log_user=list(set(a))
    
#     users = list(User.objects.all())
#     for user in users:
#         if user.id  in log_user:
#             login=log.objects.filter(user_id=user.id).first().login_time
#             logout=log.objects.filter(user_id=user.id).last().logout_time
#             attendance=Attendance.objects.create(login_time=login,logout_time=logout,user_id=user)
#             attendance.save() 
#         else:
#             continue
#     return render(request)


# # Storing Absent student 
# def storing_Absent(request) :
#     users =list(User.objects.all())
#     for user in users:
#         que=bool(log.objects.filter(user_id = user.id))
#         if que == False:
#             log.objects.create(user_id=user ,date=date.today())
#     return render(request)


# # Deleting log Enteries
# def deleting_log(request):
#     # log.objects.filter(logout_time=None).delete()
#     d=log.objects.filter(date=date.today()).last().delete()
#     #  Attendance.objects.all().delete()
#     # User.objects.get(id=7).delete()
#     return HttpResponse(d)


# Show all the attedance details
@login_required
@staff_member_required
def showing_users(request):
    # user_query=[]
    # logs=log.objects.all()
    # log_filter=log.objects.filter(logout_time=None)
    # for i in log_filter:
    #     user_query.append(i.user_id.id)
    users=User.objects.filter(is_staff=False)

    p = Paginator(users, 5)  
    
    page_number = request.GET.get('page')
    try: 
        page_obj = p.page(page_number) 
    except PageNotAnInteger:
    
        page_obj = p.page(1)
    except EmptyPage:
        page_obj = p.page(p.num_pages)
        

    return render(request,'admin.html',{'users':users, 'page_obj':page_obj})#'logs':logs,'user_query':user_query
   
    # return showing_users(request)


def user_Attendance(request,id): 
    lis=[]
    users=User.objects.all()

    
    logs=log.objects.filter(user_id=id)
    for i in logs:
        lis.append(i.user_id.id)
    user=list(set(lis))


    datef=request.GET.get('datef')
    datet=request.GET.get('datet')
    if datef is not None and datet is not None:
        dic = log.objects.filter(date__lte=datet,date__gte=datef ,user_id=id)
        
        return render(request,'user.html',{'dic':dic,'user':user})

    elif datef is not None or datet is not None:
        return HttpResponse("please select both dates")
    else:
    
        p = Paginator(logs, 5)  
    
        page_number = request.GET.get('page')
        try:
            page_obj = p.page(page_number) 
        except PageNotAnInteger:
        
            page_obj = p.page(1)
        except EmptyPage:
            page_obj = p.page(p.num_pages)
            

        return render(request,'user.html',{'logs':logs,'user':user,'page_obj':page_obj})

def All_Attendance(request):  
    date=request.GET.get('date')
    
    status=request.GET.get('status')
    if date=='':
        if  date=='' and status=='select':
            return redirect('All-Attendance')
        elif date=='' and status is not None:
            sta=log.objects.filter(status=status)
            return render(request,'All_Attendance.html',{'sta':sta})


    elif date is not None:
        if status=="select":
            filter_date=log.objects.filter(date=date )
            return render(request,'All_Attendance.html',{'filter_date':filter_date})
        else:
            filter_ds=log.objects.filter(date=date ,status=status)
            return render(request,'All_Attendance.html',{'filter_ds':filter_ds})


    else:
        all=log.objects.all().order_by('-date')

        p = Paginator(all, 10)  
    
        page_numbers = request.GET.get('page')
        try:
            page_obj = p.page(page_numbers) 
        except PageNotAnInteger:
        
            page_obj = p.page(1)
        except EmptyPage:
            page_obj = p.page(p.num_pages)

        return render(request,'All_Attendance.html',{'all':all,'page_obj':page_obj})



def search(request):
    name=request.GET.get('username')
  
    student=User.objects.filter(username__icontains=name,is_staff=False)
  
 
    return render(request,'search.html',{'student':student})



#openpyxl
def Excel_store(request,id):
    Users=User.objects.get(id=id)
    response = HttpResponse(content_type='aplication/vnd.ms-excel')
    response['content-Disposition'] = f'attachment ; filename="{Users.username}Attendance-store.xlsx"'
    row_num1=0
    wb = Workbook()
    sheet=Workbook.active
   
    rows= log.objects.filter(user_id=id).values_list( 'date','status','login_time','logout_time','working_hour').order_by('-date')
  
    # grab the active worksheet
    ws = wb.active
  
    # Data can be assigned directly to cells
    ws.append(['date','status','login_time','logout_time','working_hour'])
    # Rows can also be appended
    for row in rows:
        row_num1 += 1
        z=[i for i in row]
        ws.append(z)

    for row_range in range(2,row_num1+2):
        cell_title =ws.cell(row=row_range,column=2)
        if cell_title.value=="Absent":
            cell_title.fill =PatternFill(start_color='EE1111',end_color='EE1111',fill_type="solid")
        
    
    wb.save(response)

    return response


    
# def open_excel(request):
#     rows= log.objects.filter(user_id=3).union()
#     for i in rows:
#         print(i.login_time)
#     logs=log.objects.filter(user_id=id)
#     attendance=log.objects.filter(user_id=id)
#     # user=log.objects.all()
  
#     response = HttpResponse(content_type='aplication/ms-excel')
#     response['content-Disposition'] = 'attachment ; filename="users.xls"'

#     wb = xlwt.Workbook()
#     ws = wb.add_sheet('Users')
#     row_num = 0
#     row_num1=0
#     row_num2=0

#     font_style= xlwt.easyxf(
#     'font: name Times New Roman, colour_index red')

#     colums=['date','status','login_time','logout_time','working_hour','user-name']
#     for colum in range(len(colums)):
#         ws.write(row_num,colum,colums[colum],font_style)
           
#     style = xlwt.XFStyle()
#     style.num_format_str= 'D-MMM-YY'
#     rows= log.objects.filter(user_id=id).values_list('date','status')
#     for row in rows:
        
#         row_num += 1
#         for col_num in range(len(row)):

#             ws.write(row_num,col_num,row[col_num],style)
# #login time logout time
#     style = xlwt.XFStyle()
#     style.num_format_str= 'h:mm:sss'
#     rows= log.objects.filter(user_id=id).values_list('login_time','logout_time','working_hour')
#     for row in rows:
#         row_num1 += 1
#         for col_num in range(len(row)):
#             ws.write(row_num1,2+col_num,row[col_num],style)

# #username
#     style = xlwt.XFStyle()
    
#     rows= User.objects.filter(id=id).values_list('username')
#     for row in rows:
#         print(row)
        
#         row_num2 += 1
#         for col_num in range(len(row)):
#             ws.write(row_num2,5+col_num,row[col_num],style)
#     wb.save(response)
    
       

    # return HttpResponse("ok")

    
    